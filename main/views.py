from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .models import User, StudentProfile, Workshop, Department, Recommendation, DepartmentWeeklySlot, ApplicationDocument, Application, Notification, Workshop, WorkshopRegistration
from django.contrib.auth.decorators import login_required
from datetime import date, datetime
from .decision_tree import get_recommendation
from .models import User, StudentProfile, StaffProfile, Workshop, Department, Recommendation, Application, ApplicationDocument, DepartmentWeeklySlot, Notification, WorkshopRegistration, VolunteeringDocument, AttendanceSheet, WorkshopAttendanceSheet, Certificate
from .email_service import (
    send_application_submitted_email,
    send_application_approved_email,
    send_application_rejected_email,
    send_workshop_registration_email,
)
from functools import wraps
from django.shortcuts import get_object_or_404
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse



def landing_page(request):
    if request.user.is_authenticated:
        if request.user.role == 'staff':
            return redirect('staff_dashboard')
        return redirect('dashboard')
    return render(request, 'landing.html')

    
def auth_page(request):
    if request.user.is_authenticated:
        if request.user.role == 'staff':
            return redirect('staff_dashboard')
        return redirect('dashboard')
    return render(request, 'auth.html')

def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        try:
            user_obj = User.objects.get(email=email)
            user = authenticate(request, username=user_obj.username, password=password)
            if user is not None:
                login(request, user)
                if user_obj.role == 'staff':
                    return redirect('staff_dashboard')
                return redirect('dashboard')
            else:
                messages.error(request, 'Invalid password.')
        except User.DoesNotExist:
            messages.error(request, 'No account found with this email.')
    return redirect('auth_page')



def signup_view(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        academic_level = request.POST.get('academic_level')
        role = request.POST.get('role', 'student')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'An account with this email already exists.')
            return redirect('auth_page')

        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            role=role
        )
        StudentProfile.objects.create(
            user=user,
            full_name=full_name,
            academic_level=academic_level,
            institution='',
            phone=''
        )
        login(request, user)
        return redirect('dashboard')
    return redirect('auth_page')


@login_required
def profile_view(request):
    profile = request.user.student_profile

    if request.method == 'POST':
        profile.full_name = request.POST.get('full_name')
        profile.date_of_birth = request.POST.get('date_of_birth') or None
        profile.phone = request.POST.get('phone')
        profile.qid = request.POST.get('qid')
        profile.qid_expiry_date = request.POST.get('qid_expiry_date') or None
        profile.academic_level = request.POST.get('academic_level') or None
        profile.grade_year = request.POST.get('grade_year')
        profile.institution = request.POST.get('institution')
        profile.profile_picture = request.FILES.get('profile_picture') or None

        if request.FILES.get('id_document'):
            profile.id_document = request.FILES['id_document']

        profile.save()
        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')

    return render(request, 'profile.html', {'profile': profile})


@login_required
def mark_all_read(request):
    request.user.notifications.filter(is_read=False).update(is_read=True)
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


@login_required
def dashboard(request):
    try:
        profile = request.user.student_profile
        applications = profile.applications.all().order_by('-submitted_at')
        certificates = profile.certificates.all().order_by('-issued_at')
    except:
        profile = None
        applications = []
        certificates = []

    return render(request, 'dashboard.html', {
        'profile': profile,
        'applications': applications,
        'certificates': certificates,
    })

@login_required
def activities(request):
    profile = request.user.student_profile

    hospital_applications = profile.applications.all().order_by('-submitted_at')
    workshop_registrations = profile.workshop_registrations.all().order_by('-registered_at')
    certificates = profile.certificates.all().order_by('-issued_at')
    total_hours = sum(cert.hours for cert in certificates)

    return render(request, 'activities.html', {
        'hospital_applications': hospital_applications,
        'workshop_registrations': workshop_registrations,
        'certificates': certificates,
        'total_hours': total_hours,
    })


@login_required
def hospital_volunteering(request):
    highschool_depts = Department.objects.filter(
        is_active=True,
        eligibility__in=['highschool']
    )
    undergraduate_depts = Department.objects.filter(
        is_active=True,
        eligibility__in=['undergraduate']
    )
    return render(request, 'hospital_volunteering.html', {
        'highschool_depts': highschool_depts,
        'undergraduate_depts': undergraduate_depts,
    })

@login_required
def apply(request):
    profile = request.user.student_profile
    academic_level = profile.academic_level

    # filter departments based on academic level
    if academic_level == 'high-school':
        departments = Department.objects.filter(
            eligibility='highschool',
            is_active=True
        )
    else:
        departments = Department.objects.filter(
            is_active=True
        )

    if request.method == 'POST':
        first_choice_id = request.POST.get('first_choice_dept')
        second_choice_id = request.POST.get('second_choice_dept')
        preferred_week = request.POST.get('preferred_week')

        # validate that first choice is selected
        if not first_choice_id:
            messages.error(request, 'Please select a first choice department.')
            return redirect('apply')

        # validate preferred week is a Sunday
        if preferred_week:
            week_date = date.fromisoformat(preferred_week)
            if week_date.weekday() != 6:
                messages.error(request, 'Please select a Sunday as your week start date.')
                return redirect('apply')
        else:
            messages.error(request, 'Please select a preferred week.')
            return redirect('apply')

        # check if student already has any application for the same week
        same_week_application = Application.objects.filter(
            student=profile,
            preferred_slot__week_start_date=week_date,
            status__in=['submitted', 'approved']
        ).first()

        if same_week_application:
            messages.error(
                request,
                f'You already have an application for the week of {week_date.strftime("%B %d, %Y")}. You cannot submit two applications for the same week.'
            )
            return redirect('apply')


        # check if student already has an active approved placement
        active_approved = Application.objects.filter(
            student=profile,
            status='approved'
        ).exclude(
            approved_department__isnull=True
        ).first()

        if active_approved:
            messages.error(
                request,
                f'You already have an active approved placement at {active_approved.approved_department.name}. You cannot apply while one is still active.'
            )
            return redirect('dashboard')
            
        # get or create the weekly slot
        first_dept = Department.objects.get(id=first_choice_id)
        slot, created = DepartmentWeeklySlot.objects.get_or_create(
            department=first_dept,
            week_start_date=week_date,
            defaults={
                'filled_slots': 0
            }
        )

        # check if slot is full
        if slot.is_full():
            messages.error(request, 'Sorry, this department is full for the selected week. Please choose another week or department.')
            return redirect('apply')

        # create the application
        application = Application.objects.create(
            student=profile,
            first_choice_dept=first_dept,
            second_choice_dept=Department.objects.get(id=second_choice_id) if second_choice_id else None,
            preferred_slot=slot,
            status='submitted'
        )


        staff_users = User.objects.filter(role='staff')
        for staff in staff_users:
            Notification.objects.create(
                user=staff,
                message=f'New application received from {request.user.student_profile.full_name} for {first_dept.name}.',
                type='general'
            )

        # handle additional documents
        for file in request.FILES.getlist('extra_documents'):
            ApplicationDocument.objects.create(
                application=application,
                file=file,
                document_type='additional'
            )

        # Checking of the profile is complete
        if not profile.phone or not profile.qid or not profile.qid_expiry_date or not profile.academic_level or not profile.grade_year or not profile.institution or not profile.id_document:
            messages.error(request, 'Please complete your profile before applying.')
            return redirect('profile')
        

        # create notification for student
        Notification.objects.create(
            user=request.user,
            message=f'Your application to {first_dept.name} has been submitted successfully and is under review.',
            type='application_submitted'
        )

        messages.success(request, 'Your application has been submitted successfully!')
        return redirect('dashboard')

        

    return render(request, 'apply.html', {
        'profile': profile,
        'departments': departments,
    })

@login_required
def recommendation(request):
    return render(request, 'recommendation.html')


@login_required
def recommendation(request):
    recommended_dept = None

    if request.method == 'POST':
        answers = {
            'willing_to_travel': request.POST.get('willing_to_travel'),
            'prior_experience': request.POST.get('prior_experience'),
            'stress_tolerance': request.POST.get('stress_tolerance'),
            'patient_interaction': request.POST.get('patient_interaction'),
            'preference': request.POST.get('preference'),
            'emergency_interest': request.POST.get('emergency_interest'),
            'work_type': request.POST.get('work_type'),
        }

        student_area = request.POST.get('area', 'doha')
        academic_level = request.user.student_profile.academic_level

        recommended_dept = get_recommendation(
            answers, academic_level, student_area
        )

        # Save recommendation to database
        Recommendation.objects.create(
            student=request.user.student_profile,
            suggested_department=recommended_dept,
            answers=answers
        )

    return render(request, 'recommendation.html', {
        'recommended_dept': recommended_dept,
    })

@login_required
def register_workshop(request, workshop_id):
    if request.method == 'POST':
        workshop = Workshop.objects.get(id=workshop_id)
        profile = request.user.student_profile

        already_registered = WorkshopRegistration.objects.filter(
            student=profile,
            workshop=workshop,
            status='registered'
        ).exists()

        if already_registered:
            messages.error(request, 'You are already registered for this workshop.')
            return redirect('workshops')

        if workshop.capacity <= 0:
            messages.error(request, 'Sorry this workshop is full.')
            return redirect('workshops')

        WorkshopRegistration.objects.create(
            student=profile,
            workshop=workshop,
            status='registered'
        )

        workshop.capacity -= 1
        workshop.save()

        Notification.objects.create(
            user=request.user,
            message=f'You have successfully registered for {workshop.name}.',
            type='general'
        )

        messages.success(request, f'You have successfully registered for {workshop.name}!')
        return redirect('workshops')


# STAFF VIEWS



def staff_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'staff':
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@staff_required
def staff_dashboard(request):
    status_filter = request.GET.get('status', 'all')
    department_filter = request.GET.get('department', 'all')
    week_filter = request.GET.get('week', 'all')
    level_filter = request.GET.get('level', 'all')

    applications = Application.objects.all().order_by('-submitted_at')

    if status_filter != 'all':
        applications = applications.filter(status=status_filter)

    if department_filter != 'all':
        applications = applications.filter(
            first_choice_dept_id=department_filter
        )

    if week_filter and week_filter != 'all':
        for fmt in ['%Y-%m-%d', '%B %d, %Y', '%b %d, %Y', '%April %d, %Y']:
            try:
                week_date = datetime.strptime(week_filter.strip(), fmt).date()
                applications = applications.filter(
                    preferred_slot__week_start_date=week_date
                )
                break
            except ValueError:
                continue

    if level_filter != 'all':
        applications = applications.filter(
            student__academic_level=level_filter
        )

    total = Application.objects.count()
    pending = Application.objects.filter(status='submitted').count()
    approved = Application.objects.filter(status='approved').count()
    rejected = Application.objects.filter(status='rejected').count()

    departments = Department.objects.filter(is_active=True).order_by('name')
    weeks = DepartmentWeeklySlot.objects.values_list(
        'week_start_date', flat=True
    ).distinct().order_by('week_start_date')

    return render(request, 'staff/staff_dashboard.html', {
        'applications': applications,
        'status_filter': status_filter,
        'department_filter': department_filter,
        'week_filter': week_filter,
        'level_filter': level_filter,
        'total': total,
        'pending': pending,
        'approved': approved,
        'rejected': rejected,
        'departments': departments,
        'weeks': weeks,
    })

@login_required
@staff_required
def staff_application_detail(request, app_id):
    application = get_object_or_404(Application, id=app_id)
    return render(request, 'staff/application_detail.html', {
        'application': application,
    })

@login_required
@staff_required
def approve_application(request, app_id):
    if request.method == 'POST':
        application = get_object_or_404(Application, id=app_id)
        choice = request.POST.get('choice', 'first')

        if choice == 'second' and application.second_choice_dept:
            approved_dept = application.second_choice_dept

            if application.preferred_slot:
                week_date = application.preferred_slot.week_start_date
                slot, created = DepartmentWeeklySlot.objects.get_or_create(
                    department=approved_dept,
                    week_start_date=week_date,
                    defaults={'filled_slots': 0}
                )
                if slot.is_full():
                    messages.error(request, f'{approved_dept.name} is also full for this week.')
                    return redirect('staff_application_detail', app_id=app_id)
                application.status = 'approved'
                application.approved_department = approved_dept
                application.save()
                slot.filled_slots += 1
                slot.save()
            else:
                application.status = 'approved'
                application.approved_department = approved_dept
                application.save()

        else:
            approved_dept = application.first_choice_dept
            slot = application.preferred_slot

            if slot and slot.is_full():
                messages.error(
                    request,
                    f'{approved_dept.name} is full for this week. Please approve for second choice instead.'
                )
                return redirect('staff_application_detail', app_id=app_id)

            application.status = 'approved'
            application.approved_department = approved_dept
            application.save()

            if slot:
                slot.filled_slots += 1
                slot.save()

        Notification.objects.create(
            user=application.student.user,
            message=f'Congratulations! Your application has been approved. You will be volunteering at {approved_dept.name}.',
            type='application_approved'
        )

        try:
            send_application_approved_email(application.student.user, approved_dept)
        except Exception:
            pass

        messages.success(request, f'Application approved for {approved_dept.name}.')
        return redirect('staff_dashboard')




@login_required
@staff_required
def reject_application(request, app_id):
    if request.method == 'POST':
        application = get_object_or_404(Application, id=app_id)
        reason = request.POST.get('reason', 'No reason provided.')
        application.status = 'rejected'
        application.save()

        Notification.objects.create(
            user=application.student.user,
            message=f'Your application to {application.first_choice_dept.name} has been rejected. Reason: {reason}',
            type='application_rejected'
        )

        messages.success(request, f'Application by {application.student.full_name} has been rejected.')
        return redirect('staff_dashboard')


@login_required
@staff_required
def staff_workshops(request):
    workshops = Workshop.objects.all().order_by('-start_date')
    return render(request, 'staff/staff_workshops.html', {
        'workshops': workshops,
    })

@login_required
@staff_required
def staff_workshop_add(request):
    if request.method == 'POST':
        Workshop.objects.create(
            name=request.POST.get('name'),
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date'),
            timings=request.POST.get('timings'),
            description=request.POST.get('description'),
            capacity=request.POST.get('capacity'),
            picture=request.FILES.get('picture'),
            poster=request.FILES.get('poster'),
        )
        messages.success(request, 'Workshop added successfully!')
        return redirect('staff_workshops')
    return render(request, 'staff/staff_workshop_form.html', {
        'action': 'Add',
        'workshop': None,
    })

@login_required
@staff_required
def staff_workshop_edit(request, workshop_id):
    workshop = get_object_or_404(Workshop, id=workshop_id)
    if request.method == 'POST':
        workshop.name = request.POST.get('name')
        workshop.start_date = request.POST.get('start_date')
        workshop.end_date = request.POST.get('end_date')
        workshop.timings = request.POST.get('timings')
        workshop.description = request.POST.get('description')
        workshop.capacity = request.POST.get('capacity')
        if request.FILES.get('picture'):
            workshop.picture = request.FILES.get('picture')
        if request.FILES.get('poster'):
            workshop.poster = request.FILES.get('poster')
        workshop.save()
        messages.success(request, 'Workshop updated successfully!')
        return redirect('staff_workshops')
    return render(request, 'staff/staff_workshop_form.html', {
        'action': 'Edit',
        'workshop': workshop,
    })

@login_required
@staff_required
def staff_workshop_delete(request, workshop_id):
    if request.method == 'POST':
        workshop = get_object_or_404(Workshop, id=workshop_id)
        workshop.delete()
        messages.success(request, 'Workshop deleted successfully!')
    return redirect('staff_workshops')

@login_required
@staff_required
def staff_departments(request):
    departments = Department.objects.all().order_by('name')
    return render(request, 'staff/staff_departments.html', {
        'departments': departments,
    })

@login_required
@staff_required
def staff_department_add(request):
    if request.method == 'POST':
        Department.objects.create(
            name=request.POST.get('name'),
            location=request.POST.get('location'),
            area=request.POST.get('area'),
            supervisor=request.POST.get('supervisor'),
            timings=request.POST.get('timings'),
            eligibility=request.POST.get('eligibility'),
            total_slots=request.POST.get('total_slots', 5),
            has_evening_shift=request.POST.get('has_evening_shift') == 'on',
            is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Department added successfully!')
        return redirect('staff_departments')
    return render(request, 'staff/staff_department_form.html', {
        'action': 'Add',
        'department': None,
    })

@login_required
@staff_required
def staff_department_edit(request, dept_id):
    department = get_object_or_404(Department, id=dept_id)
    if request.method == 'POST':
        department.name = request.POST.get('name')
        department.location = request.POST.get('location')
        department.area = request.POST.get('area')
        department.supervisor = request.POST.get('supervisor')
        department.timings = request.POST.get('timings')
        department.eligibility = request.POST.get('eligibility')
        department.total_slots=request.POST.get('total_slots', 5)
        department.has_evening_shift = request.POST.get('has_evening_shift') == 'on'
        department.is_active = request.POST.get('is_active') == 'on'
        department.save()
        messages.success(request, 'Department updated successfully!')
        return redirect('staff_departments')
    return render(request, 'staff/staff_department_form.html', {
        'action': 'Edit',
        'department': department,
    })

@login_required
@staff_required
def staff_department_delete(request, dept_id):
    if request.method == 'POST':
        department = get_object_or_404(Department, id=dept_id)
        department.delete()
        messages.success(request, 'Department deleted successfully!')
    return redirect('staff_departments')





@login_required
def workshops(request):
    workshops = Workshop.objects.filter(start_date__gte=date.today()).order_by('start_date')
    return render(request, 'workshops.html', {'workshops': workshops})

def logout_view(request):
    logout(request)
    return redirect('auth_page')




@login_required
@staff_required
def staff_volunteering_documents(request):
    documents = VolunteeringDocument.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        file = request.FILES.get('file')
        if name and file:
            if documents.count() >= 2:
                messages.error(request, 'You can only have 2 volunteering documents. Please delete one first.')
            else:
                VolunteeringDocument.objects.create(name=name, file=file)
                messages.success(request, 'Document uploaded successfully!')
        return redirect('staff_volunteering_documents')

    return render(request, 'staff/staff_volunteering_documents.html', {
        'documents': documents,
    })


@login_required
@staff_required
def staff_delete_volunteering_document(request, doc_id):
    if request.method == 'POST':
        doc = get_object_or_404(VolunteeringDocument, id=doc_id)
        doc.delete()
        messages.success(request, 'Document deleted successfully!')
    return redirect('staff_volunteering_documents')


@login_required
@staff_required
def staff_attendance_list(request):
    status_filter = request.GET.get('status', 'all')
    type_filter = request.GET.get('type', 'all')

    hospital_sheets = AttendanceSheet.objects.all().order_by('-uploaded_at')
    workshop_sheets = WorkshopAttendanceSheet.objects.all().order_by('-uploaded_at')

    if status_filter != 'all':
        hospital_sheets = hospital_sheets.filter(status=status_filter)
        workshop_sheets = workshop_sheets.filter(status=status_filter)

    return render(request, 'staff/staff_attendance.html', {
        'hospital_sheets': hospital_sheets,
        'workshop_sheets': workshop_sheets,
        'status_filter': status_filter,
    })


@login_required
@staff_required
def staff_verify_attendance(request, sheet_id):
    sheet = get_object_or_404(AttendanceSheet, id=sheet_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        note = request.POST.get('staff_note', '')
        if action == 'verify':
            sheet.status = 'verified'
            sheet.staff_note = note
            sheet.save()
            Notification.objects.create(
                user=sheet.application.student.user,
                message=f'Your attendance sheet for {sheet.application.first_choice_dept.name} has been verified. Your certificate will be issued soon.',
                type='general'
            )
            messages.success(request, 'Attendance sheet verified successfully!')
        elif action == 'reject':
            sheet.status = 'rejected'
            sheet.staff_note = note
            sheet.save()
            Notification.objects.create(
                user=sheet.application.student.user,
                message=f'Your attendance sheet for {sheet.application.first_choice_dept.name} was rejected. Reason: {note}',
                type='general'
            )
            messages.success(request, 'Attendance sheet rejected.')
    return redirect('staff_attendance_list')



@login_required
@staff_required
def staff_verify_workshop_attendance(request, sheet_id):
    sheet = get_object_or_404(WorkshopAttendanceSheet, id=sheet_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        note = request.POST.get('staff_note', '')
        if action == 'verify':
            sheet.status = 'verified'
            sheet.staff_note = note
            sheet.save()
            Notification.objects.create(
                user=sheet.registration.student.user,
                message=f'Your attendance sheet for {sheet.registration.workshop.name} has been verified. Your certificate will be issued soon.',
                type='general'
            )
            messages.success(request, 'Workshop attendance verified!')
        elif action == 'reject':
            sheet.status = 'rejected'
            sheet.staff_note = note
            sheet.save()
            Notification.objects.create(
                user=sheet.registration.student.user,
                message=f'Your attendance sheet for {sheet.registration.workshop.name} was rejected. Reason: {note}',
                type='general'
            )
            messages.success(request, 'Workshop attendance rejected.')
    return redirect('staff_attendance_list')



@login_required
@staff_required
def staff_certificates_list(request):
    pending_hospital = AttendanceSheet.objects.filter(
        status='verified'
    ).exclude(
        application__certificate__isnull=False
    ).order_by('-uploaded_at')

    pending_workshop = WorkshopAttendanceSheet.objects.filter(
        status='verified'
    ).exclude(
        registration__certificate__isnull=False
    ).order_by('-uploaded_at')

    issued_certificates = Certificate.objects.all().order_by('-issued_at')

    return render(request, 'staff/staff_certificates.html', {
        'pending_hospital': pending_hospital,
        'pending_workshop': pending_workshop,
        'issued_certificates': issued_certificates,
    })


@login_required
@staff_required
def staff_upload_certificate(request, app_id):
    application = get_object_or_404(Application, id=app_id)
    if request.method == 'POST':
        Certificate.objects.create(
            activity_type='hospital',
            student=application.student,
            application=application,
            department_name=request.POST.get('department_name'),
            start_date=request.POST.get('start_date'),
            hours=request.POST.get('hours'),
            certificate_file=request.FILES.get('certificate_file'),
        )
        Notification.objects.create(
            user=application.student.user,
            message=f'Your certificate for {application.first_choice_dept.name} has been issued. You can download it from your dashboard.',
            type='general'
        )
        messages.success(request, 'Certificate uploaded successfully!')
        return redirect('staff_certificates_list')
    return render(request, 'staff/staff_certificate_form.html', {
        'application': application,
        'type': 'hospital',
    })


@login_required
@staff_required
def staff_edit_certificate(request, cert_id):
    certificate = get_object_or_404(Certificate, id=cert_id)
    if request.method == 'POST':
        certificate.department_name = request.POST.get('department_name')
        certificate.start_date = request.POST.get('start_date')
        certificate.hours = request.POST.get('hours')
        if request.FILES.get('certificate_file'):
            certificate.certificate_file = request.FILES.get('certificate_file')
        certificate.save()

        Notification.objects.create(
            user=certificate.student.user,
            message=f'Your certificate for {certificate.department_name} has been updated.',
            type='general'
        )
        messages.success(request, 'Certificate updated successfully!')
        return redirect('staff_certificates_list')

    return render(request, 'staff/staff_certificate_edit.html', {
        'certificate': certificate,
    })


@login_required
@staff_required
def staff_delete_certificate(request, cert_id):
    if request.method == 'POST':
        certificate = get_object_or_404(Certificate, id=cert_id)
        student_name = certificate.student.full_name
        dept_name = certificate.department_name
        certificate.delete()
        messages.success(
            request,
            f'Certificate for {student_name} — {dept_name} deleted successfully.'
        )
    return redirect('staff_certificates_list')



@login_required
@staff_required
def staff_upload_workshop_certificate(request, reg_id):
    registration = get_object_or_404(WorkshopRegistration, id=reg_id)
    if request.method == 'POST':
        Certificate.objects.create(
            activity_type='workshop',
            student=registration.student,
            workshop_registration=registration,
            department_name=registration.workshop.name,
            start_date=request.POST.get('start_date'),
            hours=request.POST.get('hours'),
            certificate_file=request.FILES.get('certificate_file'),
        )
        Notification.objects.create(
            user=registration.student.user,
            message=f'Your certificate for {registration.workshop.name} has been issued. You can download it from your dashboard.',
            type='general'
        )
        messages.success(request, 'Certificate uploaded successfully!')
        return redirect('staff_certificates_list')
    return render(request, 'staff/staff_certificate_form.html', {
        'registration': registration,
        'type': 'workshop',
    })


@login_required
def student_upload_attendance(request, app_id):
    application = get_object_or_404(Application, id=app_id, student=request.user.student_profile)
    if hasattr(application, 'attendance_sheet'):
        messages.error(request, 'You have already uploaded an attendance sheet for this application.')
        return redirect('dashboard')
    if request.method == 'POST':
        file = request.FILES.get('attendance_file')
        if file:
            AttendanceSheet.objects.create(
                application=application,
                file=file,
            )
            Notification.objects.create(
                user=request.user,
                message=f'Your attendance sheet for {application.first_choice_dept.name} has been uploaded and is pending verification.',
                type='general'
            )
            messages.success(request, 'Attendance sheet uploaded successfully!')
            return redirect('dashboard')
    staff_users = User.objects.filter(role='staff')
    for staff in staff_users:
        Notification.objects.create(
            user=staff,
            message=f'{request.user.student_profile.full_name} uploaded an attendance sheet for {application.first_choice_dept.name}.',
            type='general'
        )
    return render(request, 'student_upload_attendance.html', {
        'application': application,
    })

@login_required
def student_upload_workshop_attendance(request, reg_id):
    registration = get_object_or_404(
        WorkshopRegistration,
        id=reg_id,
        student=request.user.student_profile
    )

    if hasattr(registration, 'workshop_attendance'):
        if registration.workshop_attendance.status != 'rejected':
            messages.error(
                request,
                'You have already uploaded an attendance sheet for this workshop.'
            )
            return redirect('activities')

    if request.method == 'POST':
        file = request.FILES.get('attendance_file')
        if file:
            if hasattr(registration, 'workshop_attendance'):
                registration.workshop_attendance.file = file
                registration.workshop_attendance.status = 'pending'
                registration.workshop_attendance.save()
            else:
                WorkshopAttendanceSheet.objects.create(
                    registration=registration,
                    file=file,
                )

            staff_users = User.objects.filter(role='staff')
            for staff in staff_users:
                Notification.objects.create(
                    user=staff,
                    message=f'{request.user.student_profile.full_name} uploaded a workshop attendance sheet for {registration.workshop.name}.',
                    type='general'
                )

            Notification.objects.create(
                user=request.user,
                message=f'Your attendance sheet for {registration.workshop.name} has been uploaded and is pending verification.',
                type='general'
            )
            messages.success(
                request,
                'Attendance sheet uploaded successfully!'
            )
            return redirect('activities')

    return render(request, 'student_upload_workshop_attendance.html', {
        'registration': registration,
    })




@login_required
@staff_required
def staff_upload_workshop_cert_from_attendance(request, sheet_id):
    sheet = get_object_or_404(WorkshopAttendanceSheet, id=sheet_id)
    registration = sheet.registration

    if request.method == 'POST':
        Certificate.objects.create(
            activity_type='workshop',
            student=registration.student,
            workshop_registration=registration,
            department_name=request.POST.get('department_name'),
            start_date=request.POST.get('start_date'),
            hours=request.POST.get('hours'),
            certificate_file=request.FILES.get('certificate_file'),
        )
        Notification.objects.create(
            user=registration.student.user,
            message=f'Your certificate for {registration.workshop.name} has been issued. You can download it from your activities page.',
            type='general'
        )
        messages.success(request, 'Workshop certificate uploaded successfully!')
        return redirect('staff_certificates_list')

    return render(request, 'staff/staff_workshop_cert_form.html', {
        'sheet': sheet,
        'registration': registration,
    })


@login_required
@staff_required
def export_approved_excel(request):
    department_id = request.GET.get('department', '').strip()
    week = request.GET.get('week', '').strip()

    try:
        applications = Application.objects.filter(
            status='approved'
        ).select_related(
            'student', 'student__user',
            'approved_department',
            'first_choice_dept',
            'preferred_slot'
        ).order_by('student__full_name')

        if department_id and department_id != 'all':
            applications = applications.filter(
                approved_department_id=department_id
            )

        if week and week not in ['all', '']:
            week_date = None
            for fmt in ['%Y-%m-%d', '%B %d, %Y', '%b %d, %Y']:
                try:
                    week_date = datetime.strptime(week, fmt).date()
                    break
                except ValueError:
                    continue
            if week_date:
                applications = applications.filter(
                    preferred_slot__week_start_date=week_date
                )

        applications = list(applications)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Approved Volunteers'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(
            start_color='0078C2',
            end_color='0078C2',
            fill_type='solid'
        )
        header_alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alt_fill = PatternFill(
            start_color='E6F1FB',
            end_color='E6F1FB',
            fill_type='solid'
        )

        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = 'HPAP Volunteering Program — Approved Volunteers List'
        title_cell.font = Font(bold=True, size=13, color='0078C2')
        title_cell.alignment = Alignment(
            horizontal='center', vertical='center'
        )
        ws.row_dimensions[1].height = 28

        row_offset = 2

        if department_id and department_id not in ['all', '']:
            try:
                dept_obj = Department.objects.get(id=department_id)
                ws.merge_cells(f'A{row_offset}:I{row_offset}')
                cell = ws[f'A{row_offset}']
                cell.value = f'Department: {dept_obj.name} — {dept_obj.location}'
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center')
                row_offset += 1
            except Department.DoesNotExist:
                pass

        if week and week not in ['all', '']:
            ws.merge_cells(f'A{row_offset}:I{row_offset}')
            cell = ws[f'A{row_offset}']
            cell.value = f'Week Starting: {week}'
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='center')
            row_offset += 1

        row_offset += 1

        headers = [
            'No.',
            'Full Name',
            'Email',
            'QID Number',
            'Phone',
            'Academic Level',
            'Institution',
            'Department',
            'Week Start Date',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(
                row=row_offset, column=col, value=header
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[row_offset].height = 20

        for row_num, app in enumerate(applications, 1):
            row = row_offset + row_num
            try:
                dept_name = (
                    app.approved_department.name
                    if app.approved_department
                    else app.first_choice_dept.name
                    if app.first_choice_dept
                    else '—'
                )
                week_val = (
                    str(app.preferred_slot.week_start_date)
                    if app.preferred_slot
                    else '—'
                )
                data = [
                    row_num,
                    app.student.full_name or '—',
                    app.student.user.email or '—',
                    app.student.qid or '—',
                    app.student.phone or '—',
                    app.student.academic_level or '—',
                    app.student.institution or '—',
                    dept_name,
                    week_val,
                ]
            except Exception:
                continue

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if row_num % 2 == 0:
                    cell.fill = alt_fill

            ws.row_dimensions[row].height = 18

        total_row = row_offset + len(applications) + 2
        total_cell = ws.cell(
            row=total_row,
            column=1,
            value=f'Total: {len(applications)} volunteer(s)'
        )
        total_cell.font = Font(bold=True)

        column_widths = [6, 28, 28, 18, 16, 18, 28, 28, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width

        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-'
                'officedocument.spreadsheetml.sheet'
            )
        )

        dept_name_clean = 'all'
        if department_id and department_id not in ['all', '']:
            try:
                dept_name_clean = Department.objects.get(
                    id=department_id
                ).name.replace(' ', '_')
            except Department.DoesNotExist:
                pass

        week_clean = week.replace(' ', '_') if week else 'all'
        filename = f'approved_volunteers_{dept_name_clean}_{week_clean}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Export failed: {str(e)}')
        return redirect('staff_dashboard')

@login_required
@staff_required
def staff_students(request):
    search = request.GET.get('search', '')
    level_filter = request.GET.get('level', 'all')

    students = StudentProfile.objects.all().order_by('full_name')

    if search:
        students = students.filter(qid__icontains=search)

    if level_filter != 'all':
        students = students.filter(academic_level=level_filter)

    student_data = []
    for student in students:
        total_applications = student.applications.count()
        approved_applications = student.applications.filter(
            status='approved'
        ).count()
        total_certificates = student.certificates.count()
        total_hours = sum(
            cert.hours for cert in student.certificates.all()
        )
        student_data.append({
            'student': student,
            'total_applications': total_applications,
            'approved_applications': approved_applications,
            'total_certificates': total_certificates,
            'total_hours': total_hours,
        })

    return render(request, 'staff/staff_students.html', {
        'student_data': student_data,
        'search': search,
        'level_filter': level_filter,
        'total_students': students.count(),
    })


@login_required
@staff_required
def staff_student_detail(request, student_id):
    student = get_object_or_404(StudentProfile, id=student_id)
    applications = student.applications.all().order_by('-submitted_at')
    certificates = student.certificates.all().order_by('-issued_at')
    workshop_registrations = student.workshop_registrations.all().order_by('-registered_at')
    total_hours = sum(cert.hours for cert in certificates)

    return render(request, 'staff/staff_student_detail.html', {
        'student': student,
        'applications': applications,
        'certificates': certificates,
        'workshop_registrations': workshop_registrations,
        'total_hours': total_hours,
    })


    def error_404(request, exception):
        return render(request, '404.html', status=404)

    def error_500(request):
        return render(request, '500.html', status=500)

    def error_403(request, exception):
        return render(request, '403.html', status=403)